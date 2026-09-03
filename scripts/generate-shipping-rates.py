#!/usr/bin/env python3
"""
Turns Gardners' shipping price sheets into src/db/seeds/shipping-rates.ts.

A one-off maintenance tool, not part of any build. Gardners reissue these
sheets a couple of times a year; when they do, re-run this and commit the
regenerated seed. Requires openpyxl (`pip install openpyxl`) and is deliberately
NOT wired into package.json, because the vendor spreadsheets are not in the
repo — they arrive by email and live wherever you saved them.

    python3 scripts/generate-shipping-rates.py \
        --international ~/Downloads/international-cdf-july-2026.xlsx \
        --out src/db/seeds/shipping-rates.ts

The UK figures are typed in below rather than parsed: they come from a PDF
whose two-column layout (price, peak price) does not survive text extraction
reliably, and there are only fourteen of them.

The script fails loudly on any country name it cannot map to an ISO code. That
is the point: a silently skipped row is a destination we would later quote a
"rest of world" fallback for, or refuse to sell to, with nothing in the logs
saying why.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - operator tooling
    sys.exit("openpyxl is required: pip install openpyxl")


# ── Country names ────────────────────────────────────────────────────────────
# Gardners' sheets use their own spelling, which is neither ISO nor consistent
# between tabs ("UNITED STATES" on one, "UNITED STATES AMERICA" on another).
# Misspellings are theirs and are reproduced exactly — "KYRGYSTAN",
# "LEICHTENSTEIN" — because matching is on the literal cell value.
NAME_TO_ISO: dict[str, str] = {
    "AFGHANISTAN": "AF", "ALBANIA": "AL", "ALGERIA": "DZ", "ANDORRA": "AD",
    "ANGOLA": "AO", "ARGENTINA": "AR", "ARMENIA": "AM", "AUSTRALIA": "AU",
    "AUSTRIA": "AT", "AZERBAIJAN": "AZ", "BAHAMAS": "BS", "BAHRAIN": "BH",
    "BANGLADESH": "BD", "BARBADOS": "BB", "BELARUS": "BY", "BELGIUM": "BE",
    "BELIZE": "BZ", "BERMUDA": "BM", "BOLIVIA": "BO", "BOSNIA": "BA",
    "BOSNIA AND HERZEGOVINA": "BA", "BOTSWANA": "BW", "BRAZIL": "BR",
    "BRUNEI DARUSSALAM": "BN", "BULGARIA": "BG", "CAMBODIA": "KH",
    "CAMEROON": "CM", "CANADA": "CA", "CAYMAN ISLANDS": "KY",
    "CENTRAL AFRICAN REPUBLIC": "CF", "CHILE": "CL", "CHINA": "CN",
    "COLOMBIA": "CO", "COSTA RICA": "CR", "CROATIA": "HR", "CYPRUS": "CY",
    "CZECHIA": "CZ", "DENMARK": "DK", "DOMINICAN REPUBLIC": "DO", "EGYPT": "EG",
    "EL SALVADOR": "SV", "ESTONIA": "EE", "FALKLAND ISLANDS": "FK",
    "FAROE ISLANDS": "FO", "FIJI": "FJ", "FINLAND": "FI", "FRANCE": "FR",
    "FRENCH GUIANA": "GF", "FRENCH POLYNESIA": "PF", "GAMBIA": "GM",
    "GEORGIA": "GE", "GERMANY": "DE", "GHANA": "GH", "GIBRALTAR": "GI",
    "GREECE": "GR", "GREENLAND": "GL", "GRENADA": "GD", "GUADELOUPE": "GP",
    "GUAM": "GU", "GUATEMALA": "GT", "HONG KONG": "HK",
    "HONG KONG SAR, CHINA": "HK", "HUNGARY": "HU", "ICELAND": "IS",
    "INDIA": "IN", "INDONESIA": "ID", "IRELAND": "IE", "ISRAEL": "IL",
    "ITALY": "IT", "IVORY COAST": "CI", "JAMAICA": "JM", "JAPAN": "JP",
    "JORDAN": "JO", "KAZAKHSTAN": "KZ", "KENYA": "KE", "KOSOVO": "XK",
    "KUWAIT": "KW", "KYRGYSTAN": "KG", "KYRGYZSTAN": "KG", "LATVIA": "LV",
    "LEBANON": "LB", "LEICHTENSTEIN": "LI", "LITHUANIA": "LT",
    "LUXEMBOURG": "LU", "MACAU CHINA": "MO", "MALAWI": "MW", "MALAYSIA": "MY",
    "MALDIVES": "MV", "MALTA": "MT", "MARTINIQUE": "MQ", "MAURITIUS": "MU",
    "MEXICO": "MX", "MOLDOVA": "MD", "MONACO": "MC", "MONTENEGRO": "ME",
    "MOROCCO": "MA", "MYANMAR": "MM", "NAMIBIA": "NA", "NETHERLANDS": "NL",
    "NEW CALEDONIA": "NC", "NEW ZEALAND": "NZ", "NIGERIA": "NG",
    "NORTH MACEDONIA": "MK", "NORWAY": "NO", "OMAN": "OM", "PAKISTAN": "PK",
    "PANAMA": "PA", "PAPUA NEW GUINEA": "PG", "PARAGUAY": "PY", "PERU": "PE",
    "PHILIPPINES": "PH", "POLAND": "PL", "PORTUGAL": "PT", "PUERTO RICO": "PR",
    "QATAR": "QA", "ROMANIA": "RO", "RUSSIA": "RU", "SAINT LUCIA": "LC",
    "SAN MARINO": "SM", "SAUDI ARABIA": "SA", "SERBIA": "RS", "SINGAPORE": "SG",
    "SLOVAKIA": "SK", "SLOVENIA": "SI", "SOUTH AFRICA": "ZA",
    "SOUTH KOREA": "KR", "SPAIN": "ES", "SRI LANKA": "LK", "SWEDEN": "SE",
    "SWITZERLAND": "CH", "TAIWAN CHINA": "TW", "TAIWAN, CHINA": "TW",
    "TANZANIA": "TZ", "THAILAND": "TH", "TRINIDAD & TOBAGO": "TT",
    "TRINIDAD AND TOBAGO": "TT", "TUNISIA": "TN", "TURKEY": "TR",
    "UGANDA": "UG", "UKRAINE": "UA", "UNITED ARAB EMIRATES": "AE",
    "UNITED STATES": "US", "UNITED STATES AMERICA": "US", "URUGUAY": "UY",
    "UZBEKISTAN": "UZ", "VIET NAM": "VN", "ZAMBIA": "ZM", "ZIMBABWE": "ZW",
}

# Rows for a *part* of a country that ISO alpha-2 cannot address separately.
# Both carry a higher price than the mainland row of the same country, so
# keeping the mainland figure is the optimistic read — flagged in the generated
# file so nobody discovers it from an invoice.
IGNORED_SUBTERRITORIES = {
    "PORTUGAL - PORTUGUESE ISLANDS",
    "SPAIN CANARY ISLANDS",
}


# ── UK, from the Royal Mail PDF (6 October 2025) ─────────────────────────────
# (max weight in grams, price pence, peak price pence). Peak runs 17.11.25 to
# 06.01.26 and applies to large letter only — the tracked services have a single
# price, which is why their peak column is None.
#
# "Standard" (001) is 2nd class, "Premium" (002) is 1st class. Gardners send a
# large letter when the parcel fits the envelope and a tracked parcel when it
# does not, so both shapes are seeded under the same service code and chosen at
# quote time by dimensions.
UK_LARGE_LETTER_STANDARD = [(100, 145, 152), (250, 191, 198), (500, 191, 198), (750, 234, 241)]
UK_PARCEL_STANDARD = [(2000, 222, None), (20000, 257, None)]
UK_LARGE_LETTER_PREMIUM = [(100, 205, 212), (250, 259, 266), (500, 259, 266), (750, 291, 298)]
UK_PARCEL_PREMIUM = [(20000, 350, None)]

UK_SOURCE = "royal-mail-2025-10-06"
UK_EFFECTIVE_FROM = "2025-10-06"


def band_grams(header: str) -> int:
    """'up to 0.25 kg' -> 250."""
    kg = float(header.lower().replace("up to", "").replace("kg", "").strip())
    return round(kg * 1000)


def read_sheet(ws) -> tuple[list[int], dict[str, list[int]]]:
    """Returns (weight bands in grams, {iso: [pence per band]})."""
    header_row = next(
        i for i, row in enumerate(ws.iter_rows(values_only=True), 1) if row[0] == "Weights"
    )
    header = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    # The untracked tab has a blank spacer column between the country and the
    # first price, so bands are located by their header text, not by position.
    columns = [(i, band_grams(h)) for i, h in enumerate(header) if isinstance(h, str) and h.startswith("up to")]

    bands = [g for _, g in columns]
    prices: dict[str, list[int]] = {}
    unmapped: list[str] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = str(row[0]).strip().upper() if row[0] else None
        if not name or name in IGNORED_SUBTERRITORIES:
            continue

        iso = NAME_TO_ISO.get(name)
        if iso is None:
            unmapped.append(name)
            continue

        values = [row[i] for i, _ in columns]
        if any(v is None for v in values):
            # A partially filled row is not a rate we can quote from; treating
            # the gaps as free postage is the failure this guards against.
            unmapped.append(f"{name} (incomplete row)")
            continue

        pence = [round(float(v) * 100) for v in values]
        if iso in prices and prices[iso] != pence:
            # Two spellings of one country in the same tab, disagreeing. Keep
            # the dearer, so a mistake here costs margin rather than money.
            prices[iso] = [max(a, b) for a, b in zip(prices[iso], pence)]
        else:
            prices[iso] = pence

    if unmapped:
        raise SystemExit(
            f"{ws.title}: {len(unmapped)} unmapped country name(s), refusing to generate a "
            f"partial table. Add them to NAME_TO_ISO: {sorted(unmapped)}"
        )

    return bands, prices


def render_group(service: str, kind: str, source: str, effective: str, bands: list[int],
                 prices: dict[str, list[int]]) -> str:
    rows = "\n".join(
        f"      {iso}: [{', '.join(str(p) for p in prices[iso])}],"
        for iso in sorted(prices)
    )
    return (
        "  {\n"
        f"    serviceCode: '{service}',\n"
        f"    parcelKind: '{kind}',\n"
        f"    source: '{source}',\n"
        f"    effectiveFrom: '{effective}',\n"
        f"    bandsG: [{', '.join(str(b) for b in bands)}],\n"
        "    pricePence: {\n"
        f"{rows}\n"
        "    },\n"
        "  },\n"
    )


def render_uk(service: str, kind: str, table: list[tuple[int, int, int | None]]) -> str:
    bands = [b for b, _, _ in table]
    prices = [p for _, p, _ in table]
    peaks = [k for _, _, k in table]
    peak_line = (
        f"    peakPricePence: {{ GB: [{', '.join(str(p) for p in peaks)}] }},\n"
        if any(p is not None for p in peaks)
        else ""
    )
    return (
        "  {\n"
        f"    serviceCode: '{service}',\n"
        f"    parcelKind: '{kind}',\n"
        f"    source: '{UK_SOURCE}',\n"
        f"    effectiveFrom: '{UK_EFFECTIVE_FROM}',\n"
        f"    bandsG: [{', '.join(str(b) for b in bands)}],\n"
        f"    pricePence: {{ GB: [{', '.join(str(p) for p in prices)}] }},\n"
        f"{peak_line}"
        "  },\n"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--international", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    # Recorded on every seeded row so a figure can be traced back to the sheet
    # it came from. Change it when a new sheet arrives, not the filename.
    parser.add_argument("--source", default="gardners-international-cdf-2026-07")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.international, data_only=True)

    tracked_bands, tracked = read_sheet(wb["Tracked DDU"])
    untracked_bands, untracked = read_sheet(wb["Untracked"])

    # "Tracked DDP" is deliberately not seeded. It is genuinely cheaper into the
    # EU and spares the buyer a customs bill, but Gardners' I12 service-code
    # table has no code for it, so we could price a DDP order and then have no
    # way to submit one. Seed it when they tell us the code.

    intl_effective = "2026-03-16"  # both tabs: "From 16th March 2026"
    intl_source = args.source

    groups = [
        render_group("011", "parcel", intl_source, intl_effective, tracked_bands, tracked),
        render_group("010", "parcel", intl_source, intl_effective, untracked_bands, untracked),
        render_uk("001", "large_letter", UK_LARGE_LETTER_STANDARD),
        render_uk("001", "parcel", UK_PARCEL_STANDARD),
        render_uk("002", "large_letter", UK_LARGE_LETTER_PREMIUM),
        render_uk("002", "parcel", UK_PARCEL_PREMIUM),
        # BFPO rides on the same 2nd class large letter price: the PDF's row is
        # headed "Second Class Post Large Letter Including BFPO".
        render_uk("015", "large_letter", UK_LARGE_LETTER_STANDARD),
    ]

    header = f'''/**
 * Gardners shipping prices, in GBP pence, by service and destination.
 *
 * GENERATED FILE — do not edit by hand. Produced by
 * scripts/generate-shipping-rates.py from the vendor price sheets, last run
 * {dt.date.today().isoformat()} against "{args.international.name}" and the
 * Royal Mail sheet dated {UK_EFFECTIVE_FROM}.
 *
 * Prices are what Gardners charge *us* to ship a parcel. They are not retail
 * rates: the fulfilment fee, the EU customs surcharge and any margin are added
 * on top at quote time — see services/commerce/shipping.
 *
 * Each band is an upper bound in grams: a parcel is priced at the first band it
 * fits inside, and one heavier than the last band cannot be quoted at all.
 *
 * Two known imprecisions, both inherited from the source and both in the
 * customer's favour rather than ours:
 *   - Gardners price the Portuguese islands and the Canary Islands above their
 *     mainlands. ISO alpha-2 cannot express that, so those rows are dropped and
 *     PT/ES carry the mainland price.
 *   - "Tracked DDP" is not seeded at all; the I12 spec has no service code for
 *     it, so we cannot submit a DDP order even though it is cheaper into the EU.
 */

export type ParcelKind = 'large_letter' | 'parcel';

export interface ShippingRateSeedGroup {{
  serviceCode: string;
  parcelKind: ParcelKind;
  source: string;
  effectiveFrom: string;
  /** Upper bound of each weight band, in grams, ascending. */
  bandsG: number[];
  /** ISO alpha-2 -> price per band, positionally aligned with bandsG. */
  pricePence: Record<string, number[]>;
  /** Peak-season price, same shape. Only the UK large letter has one. */
  peakPricePence?: Record<string, number[]>;
}}

export const SHIPPING_RATE_SEED: ShippingRateSeedGroup[] = [
'''

    args.out.write_text(header + "".join(groups) + "];\n")

    total = sum(
        len(g["prices"]) * len(g["bands"])
        for g in [
            {"prices": tracked, "bands": tracked_bands},
            {"prices": untracked, "bands": untracked_bands},
        ]
    ) + len(UK_LARGE_LETTER_STANDARD) * 2 + len(UK_PARCEL_STANDARD) + len(UK_LARGE_LETTER_PREMIUM) + len(UK_PARCEL_PREMIUM)
    print(f"Wrote {args.out} — {len(tracked)} tracked and {len(untracked)} untracked destinations, {total} rate rows.")


if __name__ == "__main__":
    main()
